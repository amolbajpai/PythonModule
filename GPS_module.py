#import speech_recognition as sr
import os
from math import cos, asin, sqrt, pi
import pandas as pd

vsr_path = '/home/amol/Reports/VSR/'
vs_path = '/home/amol/Reports/VS/'



def find_latest_trip_advance_booking_report():
    #import os


    report_list=[]
    for i in os.listdir('/home/amol/Downloads/'):
        if "TripAdvanceBookingReport_Ver1" in i:
            report_list.append(i)

    newdict={}
    for i in report_list:
        if i=='TripAdvanceBookingReport_Ver1.xlsx':
            newdict.update({0:i})
        else:
            if i[31:32].isdigit():
                number=int(i[31:32])

            if i[31:33].isdigit():
                number=int(i[31:33])

            if i[31:34].isdigit():
                number=int(i[31:34])

            newdict.update({number:i})

    updated_excel_file=newdict[max(newdict.keys())]

    print("I am using {} for report creation".format(updated_excel_file))
    return '/home/amol/Downloads/'+updated_excel_file

def find_latest_current_status_report():
    import os
    import pandas as pd
    import ayansh as gps
    import shutil
    from datetime import datetime
    path= '/home/amol/Downloads/'
    new_path = '/home/amol/Reports/Current Status Report/'
    all_csr_files = gps.getListOfFiles(new_path)

    all_files=os.listdir(path)

    csr_files=[]

    for i in all_files:
        try:
            if i.startswith('Current_Status_Report'):
                csr_files.append(i)
        except:
            pass
      
    for i in csr_files:
        df = pd.read_excel(path+i,sheet_name='REPORT INFO')
        file_date = df.iloc[6,1]
        file_date =datetime.strptime(file_date,'%d/%m/%Y %H:%M:%S')
        file_date=file_date.strftime("%Y-%m-%d %X")
        os.rename(path+i,new_path+"Current_Status_Report "+file_date+".xls")
        #shutil.move(path+file_name,new_path+file_name.split('/')[-1]) Ex.
        print('For loop ran.....')
        
    print("Renaming done ........")

    print("No of files befor moving = ",len(all_csr_files))
    all_csr_files = gps.getListOfFiles(new_path)
    print("No of files after moving = ",len(all_csr_files))
    all_csr_files.sort()
    print("Hi Amol I have selected {} as a latest updated Current Status Report".format(all_csr_files[-1]))
    return all_csr_files[-1]



def find_latest_current_status_report_old():
    import os

    report_list=[]
    for i in os.listdir('/home/amol/Downloads/'):
        if "Current_Status_Report" in i:
            report_list.append(i)

    newdict={}
    for i in report_list:
        if i=='Current_Status_Report.xls':
            newdict.update({0:i})
        else:
            if i[23:24].isdigit():
                number=int(i[23:24])

            if i[23:25].isdigit():
                number=int(i[23:25])

            if i[23:26].isdigit():
                number=int(i[23:26])

            newdict.update({number:i})


    updated_current_status_report=newdict[max(newdict.keys())]

    print('Hi Amol, I have selected "{}" file as a latest updated file '.format(updated_current_status_report))
    return os.path.join('/home/amol/Downloads/',updated_current_status_report)

def find_latest_real_time_report():
    import os

    report_list=[]
    for i in os.listdir('/home/amol/Downloads/'):
        if "Real_Time_Report" in i:
            report_list.append(i)

    newdict={}
    for i in report_list:
        if i=='Real_Time_Report.xls':
            newdict.update({0:i})
        else:
            if i[18:19].isdigit():
                number=int(i[18:19])

            if i[18:20].isdigit():
                number=int(i[18:20])

            if i[18:21].isdigit():
                number=int(i[18:21])
            try:
                newdict.update({number:i})
            except:
                pass


    updated_real_time_report=newdict[max(newdict.keys())]

    print('Hi Amol, I have selected "{}" file as a latest updated file '.format(updated_real_time_report))
    return os.path.join('/home/amol/Downloads/',updated_real_time_report)

def df_to_excel(template,worksheet,row,col,df,output):
    from openpyxl import load_workbook
    r=row
    c=col
    #version openpyxl 3.0.0
    wb = load_workbook(template)
    print("File opened ")

    sheet = wb[worksheet]

    for i in df.index:
        c=col
        for j in df.loc[i]:
            sheet.cell(row=r,column=c).value=j
            c+=1
        r+=1
    wb.save(output)
    print("Finised..............")

def GPS_Email_Report():

    import os

    os.chdir('/home/amol/Downloads')

    file_name=list(os.listdir())
    report_list=[]
    for i in os.listdir():
        if "Current_Status_Report" in i:
            report_list.append(i)

    newdict={}
    for i in report_list:
        if i=='Current_Status_Report.xls':
            newdict.update({0:i})
        else:
            if i[23:24].isdigit():
                number=int(i[23:24])

            if i[23:25].isdigit():
                number=int(i[23:25])

            if i[23:26].isdigit():
                number=int(i[23:26])

            newdict.update({number:i})



    updated_current_status_report=newdict[max(newdict.keys())]

    print('Hi Amol, I am using "{}" file for creating report'.format(updated_current_status_report))



    import pandas as pd
    df=pd.read_excel(str('/home/amol/Downloads/'+updated_current_status_report),sheet_name="Current Status Report")
    df.to_excel("temp.xlsx",index=None)


    #from openpyxl import Workbook
    from openpyxl import load_workbook

    gps = load_workbook('/home/amol/Documents/Excel Files/Email GPS Ver 5.xlsx')
    gps_ws=gps['Current Status Report']
    novire = load_workbook("/home/amol/Downloads/temp.xlsx")
    novire_ws=novire['Sheet1']
    #vhiof = load_workbook('/home/amol/Desktop/VHIOF/vehicle hold in other fleet.xlsx')
    #vhiof_ws=vhiof.active


    gps_ws.delete_cols(1,20) # to delete previous data of the first 1 to 20 columns

    for i in range(1,novire_ws.max_row+1): # novire_ws.max_column+1
        for j in range(1,novire_ws.max_column+1):
            gps_ws.cell(row=i,column=j).value=novire_ws.cell(row=i,column=j).value

    #gps_ws_pastereport=gps['PasteReport']

    #for i in range(1,vhiof_ws.max_row+1): # novire_ws.max_column+1
    #    for j in range(1,vhiof_ws.max_column+1):
    #        gps_ws_pastereport.cell(row=i,column=j).value=vhiof_ws.cell(row=i,column=j).value


    import datetime

    date=datetime.datetime.now()
    date=date.strftime(" %d %b %Y %X")
    date=str(date)

    file_name="/home/amol/Desktop/GPS Email"+date+".xlsx"

    gps.save(file_name)

    print('Your report has been created successfully\nOutput file is located on Desktop, file name is "{}"\nFull path is {}'.format(file_name.split("/")[-1],file_name))

def find_latest_vsr():
    all_vsr_files = getListOfFiles(vsr_path)
    all_vsr_files.sort()
    print("Hi Amol, I am using ",all_vsr_files[-1])
    return all_vsr_files[-1]

def find_changes_in_controling_branch():
    import ayansh as gps
    import pandas as pd
    import numpy as np
    from datetime import timedelta
    import subprocess

    path_last_vsr = gps.find_latest_vsr()

    Updated_Current_Status_Report=gps.find_latest_current_status_report()
    Basic_Info_from_CSR = pd.read_excel(Updated_Current_Status_Report,sheet_name="Current Status Report",usecols=('Vehicle','Date/Time','Location','Speed'))
    Basic_Info_from_CSR = pd.read_excel(Updated_Current_Status_Report,sheet_name="Current Status Report",usecols=('Vehicle','Controlling Branch'))
    Basic_Info_from_CSR.rename(columns= {'Controlling Branch' : 'On Novire'},inplace=True)

    vsr= pd.read_excel(path_last_vsr,sheet_name="Vehicle_Status_Register_Quick",usecols=('Vehicle No','Vehicle control Location'))
    vsr.rename(columns={'Vehicle No' :'Vehicle', 'Vehicle control Location' : 'in Varuna' },inplace=True)
    vsr=vsr[['Vehicle','in Varuna']]
    final = Basic_Info_from_CSR.merge(vsr,on='Vehicle')
    filt = final['On Novire'] == final['in Varuna']
    mismatched = final[-filt]
    mismatched.rename(columns={'On Novire': 'Old Controlling Branch', 'in Varuna' : 'Current Controlling Branch'},inplace=True)
    mismatched['SrNo']= range(1,mismatched['Vehicle'].count()+1,1)
    mismatched=mismatched[['SrNo','Vehicle','Old Controlling Branch','Current Controlling Branch']]
    gps.df_to_excel('/home/amol/Documents/Excel Files/Update Controlling Branch Template.xlsx','Sheet1',row=2,col=1,df=mismatched,output='/home/amol/Desktop/Update Controlling Branch.xlsx')
    subprocess.call(["et",'/home/amol/Desktop/Update Controlling Branch.xlsx'],shell=False)
    print("I used ",path_last_vsr)
    print("End ..........")

def TakeCommand():
    import speech_recognition as sr
    r = sr.Recognizer()
    with sr.Microphone() as source:
        print('Say something')
        audio = r.listen(source,timeout=1,phrase_time_limit=5)

    #use Google's speech recognation
    data = ''
    try:
        data = r.recognize_google(audio)
        print('You said: '+ data)
    except sr.UnkonwnValueError:
        print('Google Speech Recognition could not understand that audio, unknown error')
    except sr.RequestError as e:
        print('Request restults from Google Speech Recognitoin service error')
    return data

def getListOfFiles(dirName):
    # create a list of file and sub directories
    # names in the given directory
    listOfFile = os.listdir(dirName)
    allFiles = list()
    # Iterate over all the entries
    for entry in listOfFile:
        # Create full path
        fullPath = os.path.join(dirName, entry)
        # If entry is a directory then get the list of files in this directory
        if os.path.isdir(fullPath):
            allFiles = allFiles + getListOfFiles(fullPath)
        else:
            allFiles.append(fullPath)

    return allFiles

def update_driver_details_db():
    import ayansh as gps
    import pandas as pd

    updated_file = gps.find_latest_trip_advance_booking_report()
    df=pd.read_excel(updated_file,usecols=["Vehicle No","Manual Driver code","Driver Mobile","Mobile No","Driver Name","Branch","Zone","Request By Mobile","From City","To City","Dated"],header=2)

    #df.sort_index(ascending=False,inplace=True)
    #df.drop_duplicates(inplace=True)
    #df.drop_duplicates(subset="Vehicle No",inplace=True)
    df['Vehicle No'] = df['Vehicle No'].str.upper()
    df.drop_duplicates(subset="Vehicle No",keep='last',inplace=True)
    df_csv = pd.read_csv("/home/amol/Documents/Excel Files/Templates/Create CSV Contacts.csv")
    def change_vehice_number(veh):
        return veh[-4:]+"-"+veh

    df_csv['Name']= df['Vehicle No'].map(change_vehice_number)
    df_csv['Yomi Name'] = df['Driver Name']
    df_csv['Phone 1 - Type'] = "Mobile"
    df_csv['Phone 1 - Value'] = df['Driver Mobile']
    df_csv['Phone 2 - Type'] = "Mobile"
    df_csv['Phone 2 - Value'] = df['Mobile No']
    df_csv['Notes'] = df['Branch'] + "---" + df['Zone'] + " | " + "FROM " + df['From City'].str.title() + " TO " + df['To City'].str.title()

    df_csv.to_csv('/home/amol/Desktop/Driver Contacts.csv',index=None)
    print("CSV created ..... ")
    #Createing file /home/amol/Documents/Excel Files/Driver_Mobile_Number.xlsx
    df_db_excel = df[['Vehicle No','Driver Mobile', 'Mobile No','Request By Mobile','Manual Driver code', 'Driver Name', 'Branch', 'Zone','From City', 'To City']]
    df_db_excel.to_excel("/home/amol/Documents/Excel Files/Driver_Mobile_Number.xlsx",index=None)
    print("Excel file: /home/amol/Documents/Excel Files/Driver_Mobile_Number.xlsx\nCreated ............")

def gps_stop_enroute_report():
    import os
    import ayansh as my_gps
    import pandas as pd
    from openpyxl import load_workbook
    #from datetime import *
    import datetime

    latest_csr_file = my_gps.find_latest_current_status_report()

    df=pd.read_excel(latest_csr_file,sheet_name="Current Status Report")
    #For filling NA with NA string
    df[df['Location'].isna()] = df[df['Location'].isna()].fillna("NA")
    print("Maximum GPS Time is = {}".format(df.loc[1,'Date/Time']))

    #calculating datetime of currnet time - latest csr  
    maximum_gps_time = df.loc[1,'Date/Time']
    maximum_gps_time = datetime.datetime.strptime(maximum_gps_time,"%Y/%m/%d %H:%M")
    print("Report delay time " ,datetime.datetime.now() - maximum_gps_time)


    df.to_excel("/home/amol/Downloads/temp.xlsx",index=None)

    gps = load_workbook('/home/amol/Documents/Excel Files/Email GPS Ver 6.xlsx')
    gps_ws=gps['Current Status Report']
    novire = load_workbook("/home/amol/Downloads/temp.xlsx")
    novire_ws=novire['Sheet1']
    #vhiof = load_workbook('/home/amol/Desktop/VHIOF/vehicle hold in other fleet.xlsx')
    #vhiof_ws=vhiof.active


    gps_ws.delete_cols(1,20) # to delete previous data of the first 1 to 20 columns 

    for i in range(1,novire_ws.max_row+1): # novire_ws.max_column+1
        for j in range(1,novire_ws.max_column+1):
            gps_ws.cell(row=i,column=j).value=novire_ws.cell(row=i,column=j).value

    #gps_ws_pastereport=gps['PasteReport']

    #for i in range(1,vhiof_ws.max_row+1): # novire_ws.max_column+1
    #    for j in range(1,vhiof_ws.max_column+1):
    #        gps_ws_pastereport.cell(row=i,column=j).value=vhiof_ws.cell(row=i,column=j).value

    #New Code
    gps_report_info = gps['REPORT INFO']

    #gps_report_info.delete_cols(6,2)
    #gps_report_info.insert_cols(6,2)

    try:
        last_vsr_file = my_gps.find_latest_vsr()
    except:
        memory_file = open('/home/amol/Documents/Excel Files/Templates/program_memory_files/last_used_vsr_file.mem','r')
        last_vsr_file = memory_file.readline()
        print("Hi I am using VSR file : ",last_vsr_file)



    vsr = pd.read_excel(last_vsr_file,sheet_name='Vehicle_Status_Register_Quick',usecols=['Vehicle No','Vehicle control Location'])
    vsr = vsr[['Vehicle No','Vehicle control Location']]


    driver_number = pd.read_excel('/home/amol/Documents/Excel Files/Driver_Mobile_Number.xlsx',sheet_name='Sheet1')
    vsr = vsr.merge(driver_number,on='Vehicle No',how='outer')


    r=1
    for i in vsr.index:
        gps_report_info.cell(row=r,column=6).value=vsr.iloc[r-1,0]
        gps_report_info.cell(row=r,column=7).value=vsr.iloc[r-1,1]
        gps_report_info.cell(row=r,column=10).value=vsr.iloc[r-1,2] #for driver mobile number add in excel
        r+=1

    #New Code

    date=datetime.datetime.now()     
    date=date.strftime(" %d %b %Y %X")
    date=str(date)

    file_name="/home/amol/Desktop/GPS Email"+date+".xlsx"
    file_name = file_name.replace(" ","_")

    gps.save(file_name)

    print('Your report has been created successfully\nOutput file is located on Desktop, file name is "{}"\nFull path is {}'.format(file_name.split("/")[-1],file_name))
    #memory_file = open('/home/amol/Documents/Excel Files/Templates/program_memory_files/last_used_vsr_file.mem','w')
    #memory_file.write(last_vsr_file)
    #memory_file.close()
    with open('/home/amol/Documents/Excel Files/Templates/program_memory_files/last_used_vsr_file.mem','w',encoding = 'utf-8') as f:
        f.write(last_vsr_file)
        f.close()
    return file_name

def rename_and_move_csr_file():
    import os
    import pandas as pd
    import ayansh as gps
    #import subprocess
    from datetime import datetime
    #path = '/home/amol/Desktop/csr/'
    path= '/home/amol/Downloads/'

    all_files=os.listdir(path)

    csr_files=[]

    for i in all_files:
        try:
            if i[:21]=='Current_Status_Report':
                csr_files.append(i)
        except:
            pass


    for i in csr_files:
        df = pd.read_excel(path+i,sheet_name='REPORT INFO')
        file_date = df.iloc[6,1]
        file_date =datetime.strptime(file_date,'%d/%m/%Y %H:%M:%S')
        file_date=file_date.strftime("%Y-%m-%d %X")
        os.rename(path+i,path+"Current_Status_Report "+file_date+".xls")

    print("Renaming done ........")

    import shutil

    all_files=os.listdir(path)

    csr_files=[]

    for i in all_files:
        try:
            if i[:21]=='Current_Status_Report':
                csr_files.append(i)
        except:
            pass


    new_path = '/home/amol/Reports/Current Status Report/'
    total_files = gps.getListOfFiles(new_path)
    print("No of files befor moving = ",len(total_files))
    for file_name in csr_files:
        shutil.move(path+file_name,new_path+file_name.split('/')[-1])
    print("Filed moved in "+new_path)
    total_files = gps.getListOfFiles(new_path)
    print("No of files after moving = ",len(total_files))

def download_current_status_report_gui():
    import pyautogui
    import time
    time.sleep(1)
    import sys
    import ayansh as gps
    import os
    png_path = '/home/amol/anaconda3/lib/python3.8/site-packages/ayansh/png_for_pyautogui/csr/'
    #os.chdir('/home/amol/Documents/01 - py/CurrentStatusReport/')
    os.chdir(png_path)
    while pyautogui.locateCenterOnScreen('Google_Chrome_Icon.png',confidence=0.80) is None:
            pass
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('Google_Chrome_Icon.png',confidence=0.9),duration=0)
    pyautogui.click()
    time.sleep(1)

    while pyautogui.locateCenterOnScreen('google_chorme_detector_app_icon.png',confidence=0.99) is None:
            pass
    #pyautogui.moveTo(pyautogui.locateCenterOnScreen('google_chorme_detector_app_icon.png',confidence=0.9),duration=0)
    time.sleep(3)

    pyautogui.hotkey('ctrl','t')
    time.sleep(1)
    pyautogui.typewrite('http://www.ivts.noviretechnologies.com/IVTS/')
    #pyautogui.typewrite('https://ivts.noviretechnologies.com/IVTS/logout.do')
    
    pyautogui.press('enter')


    while pyautogui.locateCenterOnScreen('username.png',confidence=0.80) is None:
            pass
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('username.png',confidence=0.80),duration=0)
    pyautogui.click()
    pyautogui.typewrite('varuna')
    pyautogui.typewrite(['tab'])
    #while pyautogui.locateCenterOnScreen('password.png',confidence=0.99) is None:
    #        pass
    #pyautogui.moveTo(pyautogui.locateCenterOnScreen('password.png',confidence=0.9),duration=0)
    #pyautogui.click()
    pyautogui.typewrite('vil2020')
    pyautogui.typewrite(['enter'])

    while pyautogui.locateCenterOnScreen('ReportSelectionTool.png',confidence=0.99) is None:
            pass
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('ReportSelectionTool.png',confidence=0.9),duration=0)
    pyautogui.click()


    while pyautogui.locateCenterOnScreen('SelectReport.png',confidence=0.99) is None:
            pass
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('SelectReport.png',confidence=0.9),duration=0)
    pyautogui.click()

    while pyautogui.locateCenterOnScreen('CurrentStatusReport.png',confidence=0.99) is None:
            pass
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('CurrentStatusReport.png',confidence=0.9),duration=0)
    pyautogui.click()

    while pyautogui.locateCenterOnScreen('Download.png',confidence=0.99) is None:
            pass
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('Download.png',confidence=0.9),duration=0)
    pyautogui.click()

    while pyautogui.locateCenterOnScreen('downloaded_csr_excel_file.png',confidence=0.9) is None:
            pass
    #pyautogui.moveTo(pyautogui.locateCenterOnScreen('downloaded_csr_excel_file.png',confidence=0.9),duration=0)
    #pyautogui.click()
    time.sleep(2)

    pyautogui.hotkey('alt','f4')

    print("End")

def send_gps_report_email_gui():
    png_path = '/home/amol/anaconda3/lib/python3.8/site-packages/ayansh/png_for_pyautogui/send_gps_report_email/'
    send_emails_later_tag = True
    

    import os
    os.chdir(png_path)
    #os.chdir()

    # 1 - 9 : BLY + HYD + KOL + Bhalbarh
    #controling_branchs=["01-FLBWSE","02-FLTBLS","03-FLTBLE","04-FLBNSE","05-FLTBNW","06-FLBNWS","07-FLTNN","08-HYD","09-KOL"]

    #All 
    controling_branchs=["01-FLBWSE","02-FLTBLS","03-FLTBLE","04-FLBNSE","05-FLTBNW","06-FLBNWS","07-FLTNN","08-HYD","09-KOL","10-ALL-INDIA","11-EAST","12-NORTH","13-SOUTH","14-WEST"]
    #Test 
    #controling_branchs=["13-SOUTH","14-WEST"]

    #BLY
    #controling_branchs=["01-FLBWSE","02-FLTBLS","03-FLTBLE","04-FLBNSE","05-FLTBNW","06-FLBNWS"]

    #controling_branchs=["03-FLTBLE"]

    #BLY + HYD + KOL
    #controling_branchs=["01-FLBWSE","02-FLTBLS","03-FLTBLE","04-FLBNSE","05-FLTBNW","06-FLBNWS","08-HYD","09-KOL"]

    #DHR 
    #controling_branchs=["07-FLTNN","10-ALL-INDIA","11-EAST","12-NORTH","13-SOUTH","14-WEST"]

    #import random ; random.shuffle(controling_branchs); print(controling_branchs) 

    def Type_Branch_Name(i):
        while pyautogui.locateCenterOnScreen('SrNo.png',confidence=0.99) is None:
            pass
        pyautogui.moveTo(pyautogui.locateCenterOnScreen('SrNo.png',confidence=0.99),duration=0)
        pyautogui.click()
        pyautogui.typewrite(['left'], interval=0.2)
        while pyautogui.locateCenterOnScreen('list_popup.png',confidence=0.99) is None:
            pass
        pyautogui.moveTo(pyautogui.locateCenterOnScreen('list_popup.png',confidence=0.99),duration=0)
        pyautogui.click()
        pyautogui.typewrite(i, interval=0.1)
        pyautogui.typewrite(['f9'])
        pyautogui.typewrite(['enter'], interval=0.2)
        
        
    def Found_Blank():
        time.sleep(1)
        flag=str(type(pyautogui.locateCenterOnScreen('when_report_is_blank.png',confidence=0.99)))

        if flag=="<class 'pyscreeze.Point'>":
            print("Blank")
            return True
        else:
            print("Not Blank")
            return False

    def Send_Email_Later(T):
        print('Send email later funcation stated')
        #Today_s_date=datetime.datetime.today()
        future_date=datetime.datetime(Today_s_date.year, Today_s_date.month, Today_s_date.day, HH, Start_Time+T+1)
        
        future_date=str(future_date)
        while pyautogui.locateCenterOnScreen('CreateEmail.png',confidence=0.99) is None:
            pass
        pyautogui.moveTo(pyautogui.locateCenterOnScreen('CreateEmail.png',confidence=0.9),duration=0)
        pyautogui.click()
        while pyautogui.locateCenterOnScreen('ThunderbirdSendButton.png',confidence=0.8) is None:
            pass

        #pyautogui.moveTo(pyautogui.locateCenterOnScreen('ThunderbirdSendButton.png',confidence=0.9),duration=0)
        pyautogui.hotkey('ctrl', 'v')
        
        time.sleep(1)

        pyautogui.hotkey('shift','ctrl','enter')

        while pyautogui.locateCenterOnScreen('SendAtBox.png',confidence=0.99) is None:
            pass

        pyautogui.moveTo(pyautogui.locateCenterOnScreen('SendAtBox.png',confidence=0.99),duration=0)
        pyautogui.hotkey('shift','home')
        #pyautogui.typewrite(['backspace'], interval=0)
        
        pyautogui.typewrite(future_date, interval=0.05)
        while pyautogui.locateCenterOnScreen('SendAroundButton.png',confidence=0.8) is None:
            pass
        pyautogui.moveTo(pyautogui.locateCenterOnScreen('SendAroundButton.png',confidence=0.8),duration=0.1)
        pyautogui.click()
        #time.sleep(0)

    def Send_Email_Now():
        while pyautogui.locateCenterOnScreen('CreateEmail.png',confidence=0.99) is None:
            pass
        pyautogui.moveTo(pyautogui.locateCenterOnScreen('CreateEmail.png',confidence=0.9),duration=0)
        pyautogui.click()

        while pyautogui.locateCenterOnScreen('ThunderbirdSendButton.png',confidence=0.8) is None:
            pass
        pyautogui.moveTo(pyautogui.locateCenterOnScreen('ThunderbirdSendButton.png',confidence=0.9),duration=0)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(2)
        pyautogui.click()
        #time.sleep(8)
        


    import pyautogui
    import time
    import datetime

    pyautogui.FAILSAFE = True

    Today_s_date=datetime.datetime.today()

    if Today_s_date.hour>=14:
        HH=15
        Start_Time = 0
    else:
        HH=10
        Start_Time = 30



    time.sleep(10)

    for i in controling_branchs:
        
        while True:
            try:
                print("Running for ",i)
                Type_Branch_Name(i)
                time.sleep(1)
                

                if Found_Blank():
                    print('Report is blank')
                else:
                    while pyautogui.locateCenterOnScreen('DearConcern.png',confidence=0.8) is None:
                        pass
                    pyautogui.moveTo(pyautogui.locateCenterOnScreen('DearConcern.png',confidence=0.8),duration=0)
                    pyautogui.click()
                    print("After Dear Concern")
                    
                    while True:
                        try:
                            if pyautogui.locateCenterOnScreen('End.png',confidence=0.8) is None:
                                raise NameError("Eng Not fuound")
                                
                            pyautogui.moveTo(pyautogui.locateCenterOnScreen('End.png',confidence=0.9),duration=0)
                            pyautogui.click()
                            print("Eng found")
                            pyautogui.typewrite(['left'], interval=0.2)
                            #Selecting text
                            pyautogui.hotkey('shift','ctrl','up')
                            time.sleep(0.5)
                            pyautogui.hotkey('shift','ctrl','up')
                            time.sleep(0.5)
                            #Copying text
                            pyautogui.hotkey('ctrl', 'c')
                            time.sleep(0.5)
                            #Send Email
                            T=controling_branchs.index(i)
                            if send_emails_later_tag:
                                #print("Calling Send_Email_Later fun")
                                Send_Email_Later(T)
                                print('Send email later funcation end')
                                break
                            else:
                                Send_Email_Now()
                                print("Send_Email_Now")
                                break
                                
                        except:
                            pyautogui.typewrite(['pgdn'], interval=0)
                            print("Doing page down")
                            continue
                        
                        
                                                
                time.sleep(1)
                break
            except:
                continue

    #To launch Thunderbird 
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('ThunderbirdLogo.png',confidence=0.8),duration=0)
    pyautogui.click()

     
  
def on_google_search(location):
    from googlesearch import search
    query = "wiki " + location
    #for j in search(query, tld="co.in", num=1, start=0 , stop=1,  pause=2): 
    for j in search(location, num_results=10, lang='en'):
        if 'wiki' in j:
            correct_word = j.split('/')[-1]
            return correct_word.lower()

def route_geofence_validation(clipboard_data):
    import pandas as pd
    import ayansh as gps
    import datetime
    import re
    import clipboard
    import os
    import numpy as np
    import sys


    def space_remover(text_data):
        text_data = text_data.strip()
        return re.sub(' +', ' ', text_data)



    def update_route_status(problem):
        route_geo_logs = '/home/amol/Documents/Excel Files/Templates/program_memory_files/RouteGeoLogs.xlsx'
        problem_df = pd.read_excel(route_geo_logs)
        clipboard_data['Problem'] = problem
        #todays_date = datetime.date.today() # only for Date No time
        todays_date = datetime.datetime.today()
        todays_date = todays_date.strftime("%d-%m-%Y %H:%M")
        clipboard_data['Date'] = todays_date
        problem_df = problem_df.append(clipboard_data)
        problem_df.to_excel(route_geo_logs,index=None)

    def fromat_route_data(df):
        df.loc[0,'From Location'] = df.loc[0,'From Location'].title()
        df.loc[0,'To Location'] = df.loc[0,'To Location'].title()
        df.loc[0,'Route'] = df.loc[0,'Route'].lower()
        df.loc[0,'Route'] = df.loc[0,'Route'].replace('.','')

        df.loc[0,'Route'] = space_remover(df.loc[0,'Route'])

        route_city = df.loc[0,'Route'].split(' to ')
        return route_city


    def create_route_geo_file(final):

        #for creating excel file
        final = final.loc[start_index:end_index:]
        final = final[['SrNo', 'Vehicle', 'Date/Time', 'Location', 'Latitude', 'Longitude', 'Speed']]
        final['SrNo']=range(1,final['SrNo'].count()+1,1)
        target_file_name = '/home/amol/Documents/Excel Files/Geofance/output/Route Geo-fence for '+clipboard_data['From Location'][0].title()+' to '+clipboard_data['To Location'][0].title()+'.xlsx'
        final.to_excel(target_file_name,index=None)

        #Code for adding border
        import openpyxl
        my_workbook = openpyxl.load_workbook(target_file_name)
        worksheet_sheet1 = my_workbook.active
        #Define a border
        my_border = openpyxl.styles.Side(style='thin',color='000000')
        border_ready = openpyxl.styles.Border(left=my_border,right=my_border,top=my_border,bottom=my_border)
        rows = worksheet_sheet1.iter_rows(max_row=worksheet_sheet1.max_row,max_col=worksheet_sheet1.max_column)
        for row in rows:
            for cell in row:
                cell.border = border_ready

        my_workbook.save(target_file_name)
        print("File saved..............")
        #Code for adding border /

        Subject = 'Create route Geo-fence for ' + clipboard_data['From Location'][0].title() + ' to ' + clipboard_data['To Location'][0].title()

        Body = 'Dear Team%0A%0APlease create route Geo-fence for ' + clipboard_data['From Location'][0].title() + ' to ' + clipboard_data['To Location'][0].title() +'; find route detail in attached file.'

        hlink = '<a href="mailto:service.team@testing.net.in,support@autoplant.in?cc=datta@autoplant.in,subroto.roy@noviretechnologies.com,yogesh.pawar@noviretechnologies.com,kapil.verma@test.net,harish.wadhwani@test.net,support.it.dharuhera@test.net,controlroom.dharuhera@test.net,&subject='+Subject+'&body='+Body+'">Send Email</a>'
        linkpath = target_file_name[:-4]+ 'html'
        hfile = open(linkpath,'w')
        hfile.write(hlink)
        hfile.close()
        print('Email link created.....')



    route_city = fromat_route_data(clipboard_data)
    city_dictonary_file = '/home/amol/Documents/Excel Files/Templates/program_memory_files/City_Dict.xlsx'
    city_dict_df = pd.read_excel(city_dictonary_file,index_col='Word')
    route_city_validated = []
    for city in route_city:
        try:
            search_result = city_dict_df.loc[city][0]

            print(search_result)
        except:
            print("Searcing ",city)
            search_result = gps.on_google_search(city)
            print("Result   ",search_result)
            city_dict_df.reset_index(drop=False,inplace=True)
            city_dict_df = city_dict_df.append({'Word' : city, 'Correct' : search_result },ignore_index=True)
            city_dict_df.drop_duplicates('Word',inplace=True)
            city_dict_df.set_index('Word',inplace=True)

        route_city_validated.append(search_result)
    city_dict_df.to_excel(city_dictonary_file)


    ############
    #sys.exit()

    problem = ''

    To =  route_city_validated[-1]
    From =  route_city_validated[0]

    vehicle_no = clipboard_data.loc[0,'Vehicle No']


    print("Input Values")
    print("Vehicle No: ",clipboard_data.loc[0,'Vehicle No'])
    print("From: ",From)
    print("To: ",To)

    From_point = route_city_validated[0]
    To_point = route_city_validated[-1]
    rtr_file = gps.find_latest_real_time_report()

    df = pd.read_excel(rtr_file,sheet_name='Real Time Report')
    df.replace('Invalid GPS Data',"No_Data",inplace=True)
    df.dropna(subset=['Location'],inplace=True)
    #df = pd.read_excel('/home/amol/Downloads/Real_Time_Report_Of_'+vehicle_no+'.xls',sheet_name='Real Time Report')
    final = df.drop_duplicates(subset=['Location','Latitude','Longitude']).reset_index(drop=True).copy()
    #final.dropna()
    
    final['test']=final["Location"].str.lower()
    
    #not_road = final['test'].str.contains('road',na=False) Expressway
    start_index = -1
    ### Searching To Location
    exclude_locations = ['kanpur nagar']
    enroute_keywords = ['road', 'marg', 'university','express toll','expressway', 'expressways', 'highway' ,'km away from', 'bypass','out gate']
    for i in final.index:
        exclude_flag = 0
        for excl_loc in exclude_locations:
            if excl_loc in final.loc[i,'test']:
                exclude_flag = 1
                #print('Loc ',excl_loc)
                break
        if exclude_flag == 1:
            continue
        #print(final.loc[i,'test'])
        if To_point in final.loc[i,'test']:
            found_after_char = final.loc[i,'test'].find(To_point)
            temp_location = final.loc[i,'test'][found_after_char:]
            found_flag = 0
            for e_location in enroute_keywords:
                if temp_location.find(e_location) != -1:
                    found_flag = 1
                    break
                #to exclude 'KM Away From'
                elif 'km away from' in final.loc[i,'test']:
                    found_flag = 1
                    break                    

            if found_flag == 0: # then location is ok
                start_index = i
                break
            else:
                start_index = -1
    """filt = (final['test'].str.contains(To_point,na=False)) & ~(final['test'].str.contains('road',na=False)) & ~(final['test'].str.contains('expressway',na=False)) & ~(final['test'].str.contains('km away from',na=False))  # KM Away from
    s = final[filt]
    try:
        start_index = s.index[0]
    except:
        start_index = 0"""

    ### Searching From Location
    end_index = -1
    for i in final[::-1].index:
        exclude_flag = 0
        for excl_loc in exclude_locations:
            if excl_loc in final.loc[i,'test']:
                exclude_flag = 1
                #print('Loc ',excl_loc)
                break
        if exclude_flag == 1:
            continue
        #print(final.loc[i,'test'])
        if From_point in final.loc[i,'test']:
            found_after_char = final.loc[i,'test'].find(From_point)
            temp_location = final.loc[i,'test'][found_after_char:]
            found_flag = 0
            for e_location in enroute_keywords:
                if temp_location.find(e_location) != -1:
                    found_flag = 1
                    break

               #to exclude 'KM Away From'
                elif 'km away from' in final.loc[i,'test']:
                    found_flag = 1
                    break                    


            if found_flag == 0: # then location is ok
                end_index = i
                break #commenting to get last from point
            else:
                end_index = -1
    """filt = (final['test'].str.contains(From_point,na=False)) & ~(final['test'].str.contains('road',na=False)) & ~(final['test'].str.contains('expressway',na=False)) & ~(final['test'].str.contains('km away from',na=False))  # KM Away from
    e=final[filt]
    try:
        end_index = e.index[0]
    except:
        end_index = 0
    """
    #Error handling for 'To' point

    if end_index == -1 :
        update_route_status("From Location not found")
        print("From point not found")
        problem = "From point not found"

    elif start_index == -1 :
        update_route_status("To Location not found")
        print("To Location not found")
        problem = "To Location not found"

    elif start_index > end_index :
        update_route_status("Route not found")
        print("Route not found")
        problem = "Route not found"

    else:
        problem = ''
        final = final.loc[start_index:end_index:].copy()   
        for i in route_city_validated:
            flag = 0
            filt = final['test'].str.contains(i,na=False)
            for j in list(filt):
                flag = 0
                if j == True:
                    flag+=1
                    print(i.title()," - OK",end=", ")
                    break
            if flag ==0:
                print(i.title()," - N/A",end=", ")
                problem = problem + i.title()+" - N/A, "
        print("\n######################################")


        #updateing unsuccessful route in excel file
        update_route_status(problem)

        if len(problem)==0:
            create_route_geo_file(final)
            #create_email_link()
            #lauch_email_link
    #Remove Real time report         
    #os.remove(rtr_file)
    os.rename(rtr_file,'/home/amol/RTR/{} {} to {} {}.xlsx'.format(clipboard_data.loc[0,'Vehicle No'],clipboard_data.loc[0,'From Location'],clipboard_data.loc[0,'To Location'],str(datetime.datetime.today())))
    if len(problem)==0:
        return [clipboard_data.loc[0,'From Location'],clipboard_data.loc[0,'To Location']]
    else:
        return None

    

def download_real_time_report(clipboard_data):
    import pyautogui
    import time
    import pandas as pd
    pyautogui.FAILSAFE
    time.sleep(1)
    #pyautogui.hotkey('ctrl','t')
    #detect Real Time Drop Down 
    while pyautogui.locateCenterOnScreen('RealTimeReportDropDown.png',confidence=0.8) is None:
            pass
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('RealTimeReportDropDown.png',confidence=0.8),duration=0)
    pyautogui.click()
    time.sleep(0.25)
    while pyautogui.locateCenterOnScreen('option_list.png',confidence=0.8) is None:
            pass
    
    pyautogui.click()
    time.sleep(0.25)

    pyautogui.typewrite(['tab'])
    time.sleep(0.25)
    
    pyautogui.typewrite(clipboard_data.loc[0,'Departure Date'])
    for i in range(4):
        pyautogui.typewrite(['tab'])
        time.sleep(0.2)
    pyautogui.typewrite(clipboard_data.loc[0,'Stand for Unloading Date'])
    for i in range(2):
        pyautogui.typewrite(['tab'])
        time.sleep(0.05)

    while pyautogui.locateCenterOnScreen('Download_button.png',confidence=0.8) is None:
            pass
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('Download_button.png',confidence=0.8),duration=0)
    pyautogui.moveRel(xOffset=0,yOffset=-30)
    
    pyautogui.click()

    #pyautogui.typewrite(['enter'])
    #time.sleep(0.5)
    while pyautogui.locateCenterOnScreen('filter_label.png',confidence=0.8) is None:
            pass
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('filter_label.png',confidence=0.8),duration=0)
    
    pyautogui.moveRel(xOffset=30,yOffset=0)
    pyautogui.doubleClick()
    pyautogui.typewrite(['del'])
    
    pyautogui.typewrite(clipboard_data.loc[0,'Vehicle No'])
    pyautogui.moveRel(xOffset=0,yOffset=30)
    pyautogui.click()

    while pyautogui.locateCenterOnScreen('Download_button.png',confidence=0.8) is None:
            pass
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('Download_button.png',confidence=0.8),duration=0)
    pyautogui.click()
    #Download_button

    while pyautogui.locateCenterOnScreen('Detect_real_time_report_downloaded.png',confidence=0.99) is None:
            pass
    pyautogui.moveTo(pyautogui.locateCenterOnScreen('Detect_real_time_report_downloaded.png',confidence=0.99),duration=0)
    time.sleep(1.2)
    pyautogui.hotkey('alt','f4')

def add_vehicles_gui():
    import pandas as pd
    import pyautogui
    import time
    import os
    png_path = '/home/amol/anaconda3/lib/python3.8/site-packages/ayansh/png_for_pyautogui/add_vehicle_gui'
    os.chdir(png_path)
    #pyautogui.failSafeCheck()

    while pyautogui.locateCenterOnScreen('vehicle_select_open.png',confidence=0.7) is None:
        pass
    #pyautogui.alert('Start ?')
    #time.sleep(3)


    #vehicle_select_open.png


    list_of_vehicles=pd.read_clipboard(header=None)
    list_of_vehicles=list(list_of_vehicles.iloc[:,0])


    for i in list_of_vehicles:
        while True:
            try:
                
                while pyautogui.locateCenterOnScreen('enterkeyword.png',confidence=0.8) is None:
                        pass
                pyautogui.moveTo(pyautogui.locateCenterOnScreen('enterkeyword.png',confidence=0.8),duration=0)
                
                break
            except:
                continue
        pyautogui.click()
        pyautogui.typewrite(i)
        time.sleep(0.5)
        while True:
            try:
                while pyautogui.locateCenterOnScreen('chkbutton.png',confidence=0.9) is None:
                        pass
                pyautogui.moveTo(pyautogui.locateCenterOnScreen('chkbutton.png',confidence=0.9),duration=0)
                
                
                break
            except:
                continue
        pyautogui.click()

        while True:
